from openpyxl import load_workbook#pip install openpyxl
from yattag import Doc, indent#pip install yattag
from datetime import datetime
import uuid
from random import randint
import re

def random_with_N_digits(n):
    range_start = 10 ** (n - 1)
    range_end = (10 ** n) - 1
    return randint(range_start, range_end)

def generateTnxXml():
    fileName = "Final_Excel_Format_Bank_V1"
    path = fileName+".XLSX"
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    min_row = ws.min_row
    max_row = ws.max_row
    min_col = ws.min_column
    max_col = ws.max_column

    sales_row_list:list = []
    # Will print a particular row value
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=2, max_col=ws.max_column):
        row = [cell.value for cell in row]
        sales_row_list.append(row)

    print(sales_row_list)
    doc, tag, text = Doc().tagtext()

    with tag("ENVELOPE"):
        with tag("HEADER"):
            with tag("TALLYREQUEST"):
                text("Import Data")
        with tag("BODY"):
            with tag("IMPORTDATA"):
                with tag("REQUESTDESC"):
                    with tag("REPORTNAME"):
                        text("All Masters")
                    with tag("STATICVARIABLES"):
                        with tag("SVCURRENTCOMPANY"):
                            text(fileName)
                with tag("REQUESTDATA"):
                        #Tally message gets repeated
                        #Start of One Tally message start
                    for row in sales_row_list:
                        with tag("TALLYMESSAGE", ("xmlns:UDF","TallyUDF")):
                            #Start of Voucher
                            with tag("VOUCHER",
                                     ("REMOTEID", str(uuid.uuid4())),
                                     ("VCHKEY", str(uuid.uuid4())+":00000008"),
                                     ("VCHTYPE", row[7]),
                                     ("ACTION", "Create"),
                                     ("OBJVIEW", "Accounting Voucher View")
                                     ):
                                with tag("OLDAUDITENTRYIDS.LIST",
                                             ("TYPE", "Number")):
                                    with tag("OLDAUDITENTRYIDS"):
                                        text(int(-1))
                                with tag("DATE"):
                                    saleDate = row[4]
                                    #print(saleDate.strftime("%Y%m%d"))
                                    formatedDate = "NA"
                                    if(saleDate):
                                        formatedDate = saleDate.strftime("%Y%m%d")
                                    text(str(formatedDate))
                                with tag("GUID"):
                                    text(str(uuid.uuid4()))
                                with tag("NARRATION"):
                                    text(str(row[6]))
                                with tag("PARTYLEDGERNAME"):
                                    text(str(row[3]))
                                with tag("VOUCHERTYPENAME"):
                                    text(row[7])
                                with tag("VOUCHERNUMBER"):
                                    text("1")
                                with tag("CSTFORMISSUETYPE"):
                                    text("")
                                with tag("CSTFORMRECVTYPE"):
                                    text("")
                                with tag("FBTPAYMENTTYPE"):
                                    text("Default")
                                with tag("PERSISTEDVIEW"):
                                    text("Accounting Voucher View")
                                with tag("VCHENTRYMODE"):
                                    text("Double Entry")
                                with tag("EFFECTIVEDATE"):
                                    saleDate = row[4]
                                    # print(saleDate.strftime("%Y%m%d"))
                                    formatedDate = "NA"
                                    if (saleDate):
                                        formatedDate = saleDate.strftime("%Y%m%d")
                                    text(str(formatedDate))
                                with tag("HASCASHFLOW"):
                                    text("Yes")
                                with tag("ISVATDUTYPAID"):
                                    text("Yes")
                                with tag("ALTERID"):
                                    text("867")
                                with tag("ISVATDUTYPAID"):
                                    text("433")
                                with tag("VOUCHERKEY"):
                                    text(random_with_N_digits(15))
                                with tag("ALLLEDGERENTRIES.LIST"):
                                    with tag("OLDAUDITENTRYIDS.LIST",
                                             ("TYPE", "Number")):
                                        with tag("OLDAUDITENTRYIDS"):
                                            text(int(-1))
                                    with tag("LEDGERNAME"):
                                        if re.search("Receipt", row[7], re.IGNORECASE):
                                            text(row[2])
                                        elif re.search("Payment", row[7], re.IGNORECASE):
                                            text(row[0])
                                    with tag("ISPARTYLEDGER"):
                                        text("Yes")
                                    with tag("AMOUNT"):
                                        if re.search("Receipt", row[7], re.IGNORECASE):
                                            text(round(row[5], 2))
                                        elif re.search("Payment", row[7], re.IGNORECASE):
                                            text(-round(row[5], 2))
                                with tag("ALLLEDGERENTRIES.LIST"):
                                    with tag("OLDAUDITENTRYIDS.LIST",
                                             ("TYPE", "Number")):
                                        with tag("OLDAUDITENTRYIDS"):
                                            text(int(-1))
                                    with tag("LEDGERNAME"):
                                        if re.search("Receipt", row[7], re.IGNORECASE):
                                            text(row[0])
                                        elif re.search("Payment", row[7], re.IGNORECASE):
                                            text(row[2])
                                    with tag("ISDEEMEDPOSITIVE"):
                                        text("Yes")
                                    with tag("ISPARTYLEDGER"):
                                        text("Yes")
                                    with tag("ISLASTDEEMEDPOSITIVE"):
                                        text("Yes")
                                    with tag("AMOUNT"):
                                        if re.search("Receipt", row[7], re.IGNORECASE):
                                            text(-round(row[5], 2))
                                        elif re.search("Payment", row[7], re.IGNORECASE):
                                            text(round(row[5], 2))
                                    with tag("BANKALLOCATIONS.LIST"):
                                        with tag("DATE"):
                                            saleDate = row[4]
                                            # print(saleDate.strftime("%Y%m%d"))
                                            formatedDate = "NA"
                                            if (saleDate):
                                                formatedDate = saleDate.strftime("%Y%m%d")
                                            text(str(formatedDate))
                                        with tag("INSTRUMENTDATE"):
                                            saleDate = row[4]
                                            # print(saleDate.strftime("%Y%m%d"))
                                            formatedDate = "NA"
                                            if (saleDate):
                                                formatedDate = saleDate.strftime("%Y%m%d")
                                            text(str(formatedDate))
                                        with tag("NAME"):
                                            text(str(uuid.uuid4()))
                                        with tag("TRANSACTIONTYPE"):
                                            text("Cheque/DD")
                                        with tag("PAYMENTFAVOURING"):
                                            if re.search("Receipt", row[7], re.IGNORECASE):
                                                text(row[2])
                                            elif re.search("Payment", row[7], re.IGNORECASE):
                                                text(row[0])
                                        with tag("UNIQUEREFERENCENUMBER"):
                                            text(str(uuid.uuid4()))
                                        with tag("PAYMENTMODE"):
                                            text("Transacted")
                                        with tag("BANKPARTYNAME"):
                                            if re.search("Receipt", row[7], re.IGNORECASE):
                                                text(row[2])
                                            elif re.search("Payment", row[7], re.IGNORECASE):
                                                text(row[0])
                                        with tag("CHEQUEPRINTED"):
                                            text("1")
                                        with tag("AMOUNT"):
                                            if re.search("Receipt", row[7], re.IGNORECASE):
                                                text(-round(row[5], 2))
                                            elif re.search("Payment", row[7], re.IGNORECASE):
                                                text(round(row[5], 2))

                        with tag("TALLYMESSAGE", ("xmlns:UDF", "TallyUDF")):
                            with tag("COMPANY"):
                                with tag("REMOTECMPINFO.LIST", ("MERGE", "Yes")):
                                    with tag("NAME"):
                                        text(str(uuid.uuid4()))
                                    with tag("REMOTECMPNAME"):
                                        text(fileName)
                                    with tag("REMOTECMPSTATE"):
                                        text("Odisha")

    result = indent(
        doc.getvalue(),
        indentation = '    ',
        indent_text = False
    )
    print(result)

    with open("bank_transactions"+"_"+fileName+"-new-v1.xml", "w", encoding="utf-8") as f:
        f.write(result)

    return result